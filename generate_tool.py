#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
广告系列命名工具 - 重构版生成脚本 v7
- 数据源：读取 Excel => 生成 data.json + product-code-map.json + index.html
- product_code 从 data.json 移除，改为 product-code-map.json 联动过滤
- Funnel ↔ Goal 联动、紧凑布局、全英文 UI、必填验证
"""

import json, os, re

# ── 路径配置 ────────────────────────────────────────────────────────────────
EXCEL_PATH   = r"D:/serein/工作/媒介建设/ad campaign namebuilder.xlsx"
OUTPUT_DIR   = os.path.dirname(os.path.abspath(__file__))
DATA_JSON    = os.path.join(OUTPUT_DIR, "data.json")
MAP_JSON     = os.path.join(OUTPUT_DIR, "product-code-map.json")
HTML_OUTPUT  = os.path.join(OUTPUT_DIR, "index.html")

# ── Product Type → Product Code 映射 ────────────────────────────────────────
PRODUCT_CODE_MAP = {
  "all": ["all"],
  "poe": ["all", "b1200", "b400", "b500", "b800", "c1", "c2 pro", "cx410", "cx410c", "cx810",
          "d1200", "d400", "d500", "d800", "e1 outdoor poe", "e1 outdoor se poe", "fe-p",
          "nvs8-5kb4", "nvs8-5kd4", "duo 2 poe", "duo 2v poe", "duo 3 poe", "duo 3t poe",
          "duo 3v poe", "duo floodlight poe", "duo poe", "trackmix poe",
          "rlc-1010a", "rlc-1020a", "rlc-1210a", "rlc-1212a", "rlc-1220a", "rlc-1224a",
          "rlc-1240a", "rlc-410", "rlc-410s", "rlc-411", "rlc-420", "rlc-422", "rlc-423",
          "rlc-510a", "rlc-511", "rlc-520", "rlc-520a", "rlc-522", "rlc-540a",
          "rlc-810a", "rlc-811a", "rlc-812a", "rlc-81ma", "rlc-81pa", "rlc-820a",
          "rlc-822a", "rlc-823a", "rlc-823a 16x", "rlc-823s1", "rlc-823s2", "rlc-824a",
          "rlc-830a", "rlc-833a", "rlc-840a", "rlc-842a", "rlc-843a",
          "rlk16-1200b8", "rlk16-1200d8", "rlk16-410b4d4", "rlk16-410b8",
          "rlk16-520b4d4", "rlk16-800b8", "rlk16-800d8", "rlk16-800pt8",
          "rlk16-810b8", "rlk16-812b8", "rlk16-820d8", "rlk16-833d8", "rlk16-843v8",
          "rlk4-410b4", "rlk8-1200b4", "rlk8-1200d4", "rlk8-1200v4", "rlk8-1210b4",
          "rlk8-410b2d2", "rlk8-410b4", "rlk8-410b6", "rlk8-420d4", "rlk8-500v4",
          "rlk8-510b4", "rlk8-520b2d2", "rlk8-520d4", "rlk8-800b2d2", "rlk8-800b4",
          "rlk8-800b6", "rlk8-800d4", "rlk8-800pt4", "rlk8-800tm4", "rlk8-800v4",
          "rlk8-810b2d2", "rlk8-810b4", "rlk8-810b6", "rlk8-811b4", "rlk8-812b4",
          "rlk8-820d4", "rlk8-824d4", "rlk8-833d4", "rlk8-842d4", "rlk8-843v4",
          "rlk8-cx410b4", "v1200", "v500", "v800", "cx820",
          "elite xpro poe", "elite pro floodlight poe", "rlk16-811b8", "rp-pct16md", "rp-pct8mz"],
  "wifi": ["all", "b500w", "b800w", "cx410w", "e1", "e1 outdoor", "e1 outdoor cx",
           "e1 outdoor pro", "e1 outdoor s", "e1 pro", "e1 zoom", "fe-w",
           "e1 outdoor pro+hub pro", "duo 2 wifi", "duo floodlight wifi", "duo wifi",
           "lumus", "trackmix wifi", "rlc-210w", "rlc-410w", "rlc-410ws",
           "rlc-411ws", "rlc-422w", "rlc-510wa", "rlc-511w", "rlc-511wa",
           "rlc-523wa", "rlc-542wa", "rlc-810wa", "rlc-811wa", "rlc-823s1w",
           "rlc-840wa", "rlc-843wa", "rlc-843wa-c", "rlk12-500wb4", "rlk12-800wb4",
           "rlk12-800wpt4", "rlk12-800wtm2", "rlk12-800wv4", "rlk4-210wb2",
           "rlk4-210wb4", "rlk4-211wb4", "v800w", "duo 3 wifi", "elite wifi",
           "elite floodlight wifi", "trackflex floodlight wifi"],
  "battery wifi": ["all", "argus 3 ultra", "argus 3e", "argus 4", "argus 4 pro",
                   "argus eco pro", "argus eco ultra", "argus pt lite", "argus pt ultra",
                   "argus track", "argus 3 ultra+hub", "argus 3e+hub", "argus 4+hub",
                   "argus 4 pro+hub", "argus eco ultra+hub", "argus pt ultra+hub",
                   "altas pt", "altas pt ultra", "argus", "argus 2", "argus 2e",
                   "argus 3", "argus 3 pro", "argus eco", "argus pro", "argus pt",
                   "duo", "duo 2", "trackmix", "altas",
                   "home hub mini+argus pt", "altas pt ultra+hub"],
  "battery 4g": ["all", "keen ranger pt", "duo 2 lte", "duo 4g", "go", "go plus",
                 "go pt", "go pt plus", "go pt ultra", "go ranger pt", "go ultra",
                 "trackmix lte", "trackmix lte plus", "trackmix wired lte",
                 "talon pro", "trackmix lte plus 2"],
  "doorbell": ["all", "doorbell battery", "doorbell battery+hub", "doorbell poe",
               "doorbell wifi", "home hub mini+doorbell"],
  "nvr": ["all", "home hub", "home hub pro", "rln12w", "rln16-410", "rln36", "rln8-410"]
}

# ── Funnel → Goal 联动映射 ─────────────────────────────────────────────────
FUNNEL_GOAL_MAP = {
    "awareness":      ["impression", "reach"],
    "consideration":  ["click", "view", "pageview", "signup", "atc", "checkout",
                       "engagement", "dpv", "follow"],
    "conversion":     ["sales"]
}

# ── Excel 读取（可选，找不到就用内置 fallback）────────────────────────────
def read_excel_data():
    """读取 Excel 并返回各字段枚举列表，product_code 不再从 Excel 读取"""
    try:
        import openpyxl
        wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
        ws = wb.active
        print("OK: Excel read success: " + EXCEL_PATH)
        data = {}
        for col_letter, key in [
            ("B", "market"), ("C", "product_type"), ("E", "media_funnel"),
            ("F", "ad_product"), ("G", "bidding_goal"), ("H", "url"), ("I", "custom_ref")
        ]:
            col_idx = ord(col_letter) - ord("A")
            vals = set()
            for row in range(2, ws.max_row + 1):
                v = ws.cell(row=row, column=col_idx + 1).value
                if v and str(v).strip():
                    vals.add(str(v).strip())
            data[key] = sorted(vals)
        # bidding_goal 去重合并（Excel 里可能只有部分）
        all_goals = set(data.get("bidding_goal", []))
        for goals in FUNNEL_GOAL_MAP.values():
            all_goals.update(goals)
        data["bidding_goal"] = sorted(all_goals)
        data["product_code"] = []   # 不再从 Excel 读取
        return data
    except Exception as e:
        print("WARN: Excel read failed (" + str(e) + "), using fallback data")
        return get_fallback_data()

def get_fallback_data():
    return {
        "market": ["us","ca","de","uk","fr","it","es","au","eu other","eu 2b"],
        "product_type": ["all","poe","doorbell","wifi","battery wifi","battery 4g","nvr"],
        "product_code": [],
        "media_funnel": ["awareness","consideration","conversion"],
        "ad_product": ["search","ytb","discovery","dg","pmax","gdn",
                       "fb","ig","tw","bing","reddit","programmatic","amg",
                       "fb-ig","tiktok","linkedin"],
        "bidding_goal": ["impression","reach","click","view","pageview",
                         "signup","atc","checkout","sales","engagement","dpv","follow"],
        "url": ["dtc","amz","tkshop","post"],
        "custom_ref": ["rm","aon","season","promo","bundle","clearance",
                       "new arrival","refurbished","pre-order","limited edition",
                       "collab","holiday","back to school","flash sale",
                       "free shipping","warranty","app only","influencer",
                       "retargeting","lookalike"]
    }

# ── 生成 data.json ──────────────────────────────────────────────────────────
def generate_data_json(data):
    """data.json：不含 product_code（由 product-code-map.json 管理）"""
    out = {k: v for k, v in data.items() if k != "product_code"}
    with open(DATA_JSON, "w", encoding="utf-8") as f:
        json.dump(out, f, ensure_ascii=False, indent=2)
    print("OK: Generated data.json (without product_code)")

def generate_product_code_map_json():
    """生成 product-code-map.json"""
    with open(MAP_JSON, "w", encoding="utf-8") as f:
        json.dump(PRODUCT_CODE_MAP, f, ensure_ascii=False, indent=2)
    print("OK: Generated product-code-map.json")

# ── 生成 index.html ────────────────────────────────────────────────────────
def get_html_template():
    return r"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>Ad Campaign Name Builder</title>
<style>
/* ── Reset & Base ───────────────────────────────────────────────────── */
*, *::before, *::after { box-sizing: border-box; margin: 0; padding: 0; }
:root {
  --bg: #f0f2f5; --card: #fff; --primary: #4f6ef7; --primary-hover: #3b5de7;
  --danger: #e74c3c; --success: #27ae60; --text: #1a1a2e; --text2: #555;
  --border: #dde1e7; --radius: 10px; --shadow: 0 2px 12px rgba(0,0,0,.08);
}
body { font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', Roboto, sans-serif;
       background: var(--bg); color: var(--text); line-height: 1.6; padding: 20px; }
.container { max-width: 960px; margin: 0 auto; }

/* ── Header ─────────────────────────────────────────────────────────── */
.header { background: linear-gradient(135deg, #4f6ef7 0%, #6a82fb 100%);
  color: #fff; border-radius: var(--radius); padding: 28px 32px; margin-bottom: 24px;
  display: flex; align-items: center; gap: 16px; box-shadow: var(--shadow); }
.header-icon { font-size: 36px; }
.header h1 { font-size: 24px; font-weight: 700; }
.header p  { font-size: 14px; opacity: .85; margin-top: 4px; }

/* ── Card ───────────────────────────────────────────────────────────── */
.card { background: var(--card); border-radius: var(--radius); box-shadow: var(--shadow);
        padding: 24px; margin-bottom: 20px; }
.card-header { display: flex; align-items: center; gap: 10px; margin-bottom: 20px;
              padding-bottom: 14px; border-bottom: 1px solid var(--border); }
.card-icon { font-size: 22px; }
.card-title { font-size: 17px; font-weight: 600; }

/* ── Field compact row (Market + Product Type) ─────────────────────── */
.field-compact { display: contents; }
.field-compact > .form-group { margin-bottom: 0; }
.form-row { display: flex; gap: 16px; margin-bottom: 18px; }
.form-row .form-group { flex: 1; min-width: 0; }

/* ── Form elements ─────────────────────────────────────────────────── */
.form-group { margin-bottom: 18px; }
.form-group label { display: block; font-weight: 600; font-size: 13px;
                   margin-bottom: 6px; color: var(--text); }
.form-group label .req { color: var(--danger); margin-left: 2px; }
.form-hint { font-size: 11px; color: var(--text2); margin-top: 4px; }
select, input[type="text"] { width: 100%; padding: 9px 12px; border: 1.5px solid var(--border);
  border-radius: 7px; font-size: 13px; background: #fafbfc; transition: border .2s; }
select:focus, input[type="text"]:focus { outline: none; border-color: var(--primary);
  box-shadow: 0 0 0 3px rgba(79,110,247,.12); }
select[multiple] { min-height: 80px; padding: 6px; }
select[multiple] option { padding: 5px 8px; border-radius: 4px; cursor: pointer; }
select[multiple] option:checked { background: var(--primary); color: #fff; }

/* ── Tags ───────────────────────────────────────────────────────────── */
.tags { display: flex; flex-wrap: wrap; gap: 6px; margin-top: 6px; }
.tag { display: inline-flex; align-items: center; gap: 4px; padding: 4px 10px;
       background: #eef1fb; border: 1px solid #d0d7f0; border-radius: 20px;
       font-size: 12px; color: var(--primary); cursor: pointer; transition: all .2s; }
.tag:hover { background: #dce3fa; }
.tag.selected { background: var(--primary); color: #fff; border-color: var(--primary); }
.tag .remove { margin-left: 2px; font-weight: bold; opacity: .7; }
.tag .remove:hover { opacity: 1; }

/* ── Buttons ───────────────────────────────────────────────────────── */
.btn { display: inline-flex; align-items: center; gap: 6px; padding: 10px 20px;
       border: none; border-radius: 7px; font-size: 13px; font-weight: 600;
       cursor: pointer; transition: all .2s; }
.btn-primary { background: var(--primary); color: #fff; }
.btn-primary:hover { background: var(--primary-hover); transform: translateY(-1px); }
.btn-danger  { background: var(--danger); color: #fff; }
.btn-danger:hover  { background: #c0392b; }
.btn-outline { background: transparent; color: var(--text2); border: 1.5px solid var(--border); }
.btn-outline:hover { border-color: var(--primary); color: var(--primary); }
.btn-sm { padding: 6px 12px; font-size: 12px; }
.btn-group { display: flex; gap: 10px; flex-wrap: wrap; margin-top: 16px; }

/* ── Preview ────────────────────────────────────────────────────────── */
.preview-box { background: #f8f9fc; border: 1.5px dashed var(--border);
  border-radius: 8px; padding: 16px; font-size: 14px; word-break: break-all;
  min-height: 52px; display: flex; align-items: center; gap: 10px; }
.preview-box .result { font-family: 'SF Mono', 'Fira Code', monospace;
  font-size: 15px; font-weight: 600; color: var(--primary); }
.preview-actions { display: flex; gap: 8px; margin-top: 12px; }

/* ── Field breakdown ────────────────────────────────────────────────── */
.breakdown { margin-top: 16px; }
.breakdown-item { display: flex; justify-content: space-between; align-items: center;
  padding: 8px 12px; border-radius: 6px; font-size: 13px; }
.breakdown-item:nth-child(odd) { background: #f8f9fc; }
.bd-label { color: var(--text2); font-weight: 500; }
.bd-value { font-weight: 600; color: var(--text); }

/* ── Copy history ───────────────────────────────────────────────────── */
.history-list { max-height: 200px; overflow-y: auto; }
.history-item { display: flex; justify-content: space-between; align-items: center;
  padding: 8px 12px; border-bottom: 1px solid var(--border); font-size: 13px; }
.history-item:last-child { border-bottom: none; }
.history-text { font-family: monospace; color: var(--primary); cursor: pointer;
  max-width: 70%; overflow: hidden; text-overflow: ellipsis; white-space: nowrap; }
.history-text:hover { text-decoration: underline; }
.history-time { font-size: 11px; color: var(--text2); }

/* ── Custom help panel ──────────────────────────────────────────────── */
.custom-help { margin-top: 8px; border: 1px solid var(--border); border-radius: 8px;
  overflow: hidden; }
.custom-help-header { display: flex; justify-content: space-between; align-items: center;
  padding: 10px 14px; background: #f8f9fc; cursor: pointer; font-size: 13px;
  font-weight: 600; user-select: none; }
.custom-help-header:hover { background: #eef1fb; }
.custom-help-body { padding: 14px; background: #fff; }
.custom-help-body.hidden { display: none; }
.help-tags { display: flex; flex-wrap: wrap; gap: 6px; }
.help-tag { padding: 4px 10px; background: #eef1fb; border-radius: 20px;
  font-size: 12px; color: var(--primary); cursor: pointer; }
.help-tag:hover { background: var(--primary); color: #fff; }
.help-note { font-size: 11px; color: var(--text2); margin-top: 8px; }

/* ── Validation warning ────────────────────────────────────────────── */
.validation-warning { background: #fff3f0; border: 1.5px solid var(--danger);
  border-radius: 8px; padding: 12px 16px; margin-bottom: 16px; color: var(--danger);
  font-size: 13px; display: none; }
.validation-warning.show { display: flex; align-items: center; gap: 8px; }
.validation-warning ul { margin: 4px 0 0 18px; }

/* ── Responsive ─────────────────────────────────────────────────────── */
@media (max-width: 640px) {
  .form-row { flex-direction: column; gap: 0; }
  .preview-actions { flex-wrap: wrap; }
}
</style>
</head>
<body>

<div class="container">
  <!-- Header -->
  <div class="header">
    <div class="header-icon">📋</div>
    <div>
      <h1>Ad Campaign Name Builder</h1>
      <p>Generate standardized ad campaign names across Google, Microsoft & Meta platforms</p>
    </div>
  </div>

  <!-- Validation Warning -->
  <div class="validation-warning" id="validationWarning">
    <span>⚠️</span>
    <div>
      <strong>Please fill in all required fields:</strong>
      <ul id="missingFieldsList"></ul>
    </div>
  </div>

  <!-- Step 1: Select Fields -->
  <div class="card">
    <div class="card-header">
      <div class="card-icon">📝</div>
      <div>
        <div class="card-title">Select Fields</div>
        <div style="font-size:12px;color:var(--text2);margin-top:2px;">Fields marked * are required</div>
      </div>
    </div>

    <!-- Row 1: Market + Product Type (compact side-by-side) -->
    <div class="form-row">
      <div class="form-group">
        <label>Market <span class="req">*</span></label>
        <select id="market" multiple></select>
        <div class="form-hint">Hold Ctrl/Cmd to select multiple</div>
      </div>
      <div class="form-group">
        <label>Product Type <span class="req">*</span></label>
        <select id="productType"></select>
      </div>
    </div>

    <!-- Row 2: Product Code (dynamically filtered) -->
    <div class="form-group">
      <label>Product Code <span class="req">*</span></label>
      <select id="productCode" multiple></select>
      <div class="form-hint">Options filtered by Product Type selection</div>
    </div>

    <!-- Row 3: Media Funnel + Ad Product -->
    <div class="form-row">
      <div class="form-group">
        <label>Media Funnel <span class="req">*</span></label>
        <select id="mediaFunnel"></select>
      </div>
      <div class="form-group">
        <label>Ad Product <span class="req">*</span></label>
        <select id="adProduct" multiple></select>
      </div>
    </div>

    <!-- Row 4: Bidding Goal + URL -->
    <div class="form-row">
      <div class="form-group">
        <label>Bidding Goal <span class="req">*</span></label>
        <select id="biddingGoal" multiple></select>
        <div class="form-hint">Available options depend on Media Funnel</div>
      </div>
      <div class="form-group">
        <label>URL <span class="req">*</span></label>
        <select id="url" multiple></select>
      </div>
    </div>

    <!-- Row 5: Custom -->
    <div class="form-group">
      <label>Custom</label>
      <input type="text" id="custom" placeholder="e.g. rm, season, promo (lowercase+hyphens only)">
      <div class="custom-help" id="customHelp">
        <div class="custom-help-header" onclick="toggleCustomHelp()">
          <span>💡 Common Custom Values (reference)</span>
          <span id="customHelpArrow">▼</span>
        </div>
        <div class="custom-help-body" id="customHelpBody">
          <div class="help-tags" id="customHelpTags"></div>
          <div class="help-note">Click to add; click again to remove. Custom must be lowercase English + numbers + hyphens only, no underscores.</div>
        </div>
      </div>
    </div>

    <!-- Action Buttons -->
    <div class="btn-group">
      <button class="btn btn-primary" onclick="generateName()">✨ Generate Name</button>
      <button class="btn btn-outline" onclick="copyResult()">📋 Copy</button>
      <button class="btn btn-outline" onclick="copyAll()">📄 Copy All (CSV)</button>
      <button class="btn btn-danger" onclick="resetAll()">🔁 Reset All Fields</button>
    </div>
  </div>

  <!-- Step 2: Preview -->
  <div class="card">
    <div class="card-header">
      <div class="card-icon">⭐</div>
      <div>
        <div class="card-title">Preview</div>
        <div style="font-size:12px;color:var(--text2);margin-top:2px;">Generated campaign name</div>
      </div>
    </div>
    <div class="preview-box" id="previewBox">
      <span style="color:var(--text2);">Fill in required fields and click "Generate Name"</span>
    </div>
    <div class="preview-actions" id="previewActions" style="display:none;">
      <button class="btn btn-sm btn-outline" onclick="copyResult()">📋 Copy</button>
      <button class="btn btn-sm btn-outline" onclick="copyAll()">📄 Copy All (CSV)</button>
    </div>
    <div class="breakdown" id="breakdown" style="display:none;"></div>
  </div>

  <!-- Step 3: Copy History -->
  <div class="card">
    <div class="card-header">
      <div class="card-icon">📰</div>
      <div>
        <div class="card-title">Copy History</div>
        <div style="font-size:12px;color:var(--text2);margin-top:2px;">Recently generated names</div>
      </div>
    </div>
    <div class="history-list" id="historyList">
      <div style="padding:20px;text-align:center;color:var(--text2);font-size:13px;">
        No history yet
      </div>
    </div>
  </div>
</div>

<script>
/* ═══════════════════════════════════════════════════════════════════════════
   DATA & STATE
   ═════════════════════════════════════════════════════════════════════════ */
let ENUM_DATA = {};
let PRODUCT_CODE_MAP = {};
let FUNNEL_GOAL_MAP = {};
let copyHistory = [];

/* Funnel → Goal mapping (embedded fallback) */
const DEFAULT_FUNNEL_GOAL_MAP = {
  "awareness":     ["impression", "reach"],
  "consideration": ["click","view","pageview","signup","atc","checkout","engagement","dpv","follow"],
  "conversion":    ["sales"]
};

/* ═══════════════════════════════════════════════════════════════════════════
   FETCH DATA (data.json + product-code-map.json)
   ═════════════════════════════════════════════════════════════════════════ */
async function fetchData() {
  let dataOk = false, mapOk = false;

  /* 1. Fetch data.json */
  try {
    const resp = await fetch('./data.json');
    if (resp.ok) {
      ENUM_DATA = await resp.json();
      dataOk = true;
      console.log('[data.json] loaded from server');
    }
  } catch(e) { console.warn('[data.json] fetch failed', e); }

  /* 2. Fetch product-code-map.json */
  try {
    const resp = await fetch('./product-code-map.json');
    if (resp.ok) {
      PRODUCT_CODE_MAP = await resp.json();
      mapOk = true;
      console.log('[product-code-map.json] loaded from server');
    }
  } catch(e) { console.warn('[product-code-map.json] fetch failed', e); }

  /* 3. Fetch funnel-goal-map (optional) */
  try {
    const resp = await fetch('./funnel-goal-map.json');
    if (resp.ok) {
      FUNNEL_GOAL_MAP = await resp.json();
      console.log('[funnel-goal-map.json] loaded');
    }
  } catch(e) { /* ignore */ }

  /* Fallbacks if fetch failed */
  if (!dataOk) {
    ENUM_DATA = getFallbackData();
    console.log('[data.json] using embedded fallback');
  }
  if (!mapOk) {
    PRODUCT_CODE_MAP = getFallbackMap();
    console.log('[product-code-map.json] using embedded fallback');
  }
  if (Object.keys(FUNNEL_GOAL_MAP).length === 0) {
    FUNNEL_GOAL_MAP = DEFAULT_FUNNEL_GOAL_MAP;
  }

  initUI();
}

function getFallbackData() {
  return {
    "market": ["us","ca","de","uk","fr","it","es","au","eu other","eu 2b"],
    "product_type": ["all","poe","doorbell","wifi","battery wifi","battery 4g","nvr"],
    "media_funnel": ["awareness","consideration","conversion"],
    "ad_product": ["search","ytb","discovery","dg","pmax","gdn",
                   "fb","ig","tw","bing","reddit","programmatic","amg",
                   "fb-ig","tiktok","linkedin"],
    "bidding_goal": ["impression","reach","click","view","pageview",
                     "signup","atc","checkout","sales","engagement","dpv","follow"],
    "url": ["dtc","amz","tkshop","post"],
    "custom_ref": ["rm","aon","season","promo","bundle","clearance",
                   "new arrival","refurbished","pre-order","limited edition",
                   "collab","holiday","back to school","flash sale",
                   "free shipping","warranty","app only","influencer",
                   "retargeting","lookalike"]
  };
}

function getFallbackMap() {
  return ##PRODUCT_CODE_MAP_PLACEHOLDER##;
}

/* ═══════════════════════════════════════════════════════════════════════════
   INIT UI
   ═════════════════════════════════════════════════════════════════════════ */
function initUI() {
  /* Populate single-select dropdowns */
  fillSelect('market',            ENUM_DATA.market,         true);
  fillSelect('productType',       ENUM_DATA.product_type,   false);
  fillSelect('mediaFunnel',       ENUM_DATA.media_funnel,  false);
  fillSelect('adProduct',         ENUM_DATA.ad_product,     true);
  fillSelect('biddingGoal',       ENUM_DATA.bidding_goal,  true);
  fillSelect('url',               ENUM_DATA.url,            true);

  /* Product Code: initially populate from map using default productType */
  updateProductCodeOptions();

  /* Custom reference tags */
  renderCustomHelpTags();

  /* Listeners */
  document.getElementById('mediaFunnel').addEventListener('change', onFunnelChange);
  document.getElementById('productType').addEventListener('change', onProductTypeChange);

  /* Auto-generate on any change */
  document.querySelectorAll('select, input').forEach(el => {
    el.addEventListener('change', () => { if (document.getElementById('previewBox').dataset.generated) generateName(); });
    el.addEventListener('input',  () => { if (document.getElementById('previewBox').dataset.generated) generateName(); });
  });
}

function fillSelect(id, items, multiple) {
  const sel = document.getElementById(id);
  sel.innerHTML = '';
  if (!multiple) {
    const ph = document.createElement('option');
    ph.value = '';
    ph.textContent = '-- Select --';
    sel.appendChild(ph);
  }
  (items || []).forEach(v => {
    const opt = document.createElement('option');
    opt.value = v;
    opt.textContent = v;
    sel.appendChild(opt);
  });
}

/* ═══════════════════════════════════════════════════════════════════════════
   PRODUCT TYPE → PRODUCT CODE 联动
   ═════════════════════════════════════════════════════════════════════════ */
function onProductTypeChange() {
  updateProductCodeOptions();
  /* Auto re-generate if already generated */
  if (document.getElementById('previewBox').dataset.generated) generateName();
}

function updateProductCodeOptions() {
  const typeSel = document.getElementById('productType');
  const codeSel = document.getElementById('productCode');
  const selectedType = typeSel.value;

  let codes = [];
  if (selectedType && PRODUCT_CODE_MAP[selectedType]) {
    codes = PRODUCT_CODE_MAP[selectedType];
  } else {
    /* Fallback: show all codes from all types except 'all' */
    const allCodes = new Set();
    Object.values(PRODUCT_CODE_MAP).forEach(arr => arr.forEach(c => { if (c !== 'all') allCodes.add(c); }));
    codes = ['all', ...Array.from(allCodes).sort()];
  }

  /* Remember previous selection */
  const prevSelected = Array.from(codeSel.selectedOptions).map(o => o.value);
  fillSelect('productCode', codes, true);

  /* Restore previous selection if still available */
  Array.from(codeSel.options).forEach(opt => {
    if (prevSelected.includes(opt.value)) opt.selected = true;
  });
}

/* ═══════════════════════════════════════════════════════════════════════════
   FUNNEL → GOAL 联动
   ═════════════════════════════════════════════════════════════════════════ */
function onFunnelChange() {
  const funnel = document.getElementById('mediaFunnel').value;
  const goalSel = document.getElementById('biddingGoal');
  const allowed = FUNNEL_GOAL_MAP[funnel] || [];

  const prevSelected = Array.from(goalSel.selectedOptions).map(o => o.value);
  fillSelect('biddingGoal', allowed, true);

  /* Keep valid previously selected goals */
  Array.from(goalSel.options).forEach(opt => {
    if (prevSelected.includes(opt.value)) opt.selected = true;
  });

  if (document.getElementById('previewBox').dataset.generated) generateName();
}

/* ═══════════════════════════════════════════════════════════════════════════
   GENERATE
   ═════════════════════════════════════════════════════════════════════════ */
function getSelectedValues(id) {
  const el = document.getElementById(id);
  if (el.multiple) {
    return Array.from(el.selectedOptions).map(o => o.value).filter(v => v);
  }
  return el.value ? [el.value] : [];
}

function validateRequired() {
  const required = [
    { id: 'market',       name: 'Market' },
    { id: 'productType',  name: 'Product Type' },
    { id: 'productCode',  name: 'Product Code' },
    { id: 'mediaFunnel',  name: 'Media Funnel' },
    { id: 'adProduct',    name: 'Ad Product' },
    { id: 'biddingGoal',  name: 'Bidding Goal' },
    { id: 'url',          name: 'URL' }
  ];
  const missing = [];
  required.forEach(f => {
    const vals = getSelectedValues(f.id);
    if (vals.length === 0) missing.push(f.name);
  });
  return missing;
}

function generateName() {
  const missing = validateRequired();
  const warnEl = document.getElementById('validationWarning');
  const listEl = document.getElementById('missingFieldsList');

  if (missing.length > 0) {
    warnEl.classList.add('show');
    listEl.innerHTML = missing.map(m => `<li>${m}</li>`).join('');
    document.getElementById('previewBox').innerHTML =
      '<span style="color:var(--text2);">Fill in all required fields to generate a name</span>';
    document.getElementById('previewBox').dataset.generated = '';
    document.getElementById('previewActions').style.display = 'none';
    document.getElementById('breakdown').style.display = 'none';
    return;
  }
  warnEl.classList.remove('show');

  /* Build name */
  const parts = [
    ...getSelectedValues('market'),
    ...getSelectedValues('productType'),
    ...getSelectedValues('productCode'),
    ...getSelectedValues('mediaFunnel'),
    ...getSelectedValues('adProduct'),
    ...getSelectedValues('biddingGoal'),
    ...getSelectedValues('url')
  ];

  const customVal = document.getElementById('custom').value.trim().toLowerCase();
  if (customVal) {
    if (!/^[a-z0-9\- ]+$/.test(customVal) || customVal.includes('_')) {
      alert('Custom field can only contain lowercase English, numbers, hyphens and spaces. No underscores allowed.');
      return;
    }
    parts.push(customVal);
  }

  const name = parts.join('_');
  const box = document.getElementById('previewBox');
  box.innerHTML = `<span class="result">${name}</span>`;
  box.dataset.generated = '1';

  document.getElementById('previewActions').style.display = 'flex';
  renderBreakdown(parts);
  addHistory(name);
}

/* ═══════════════════════════════════════════════════════════════════════════
   BREAKDOWN
   ═════════════════════════════════════════════════════════════════════════ */
function renderBreakdown(parts) {
  const el = document.getElementById('breakdown');
  const labels = [
    ...getSelectedValues('market').map(() => 'Market'),
    ...getSelectedValues('productType').map(() => 'Product Type'),
    ...getSelectedValues('productCode').map(() => 'Product Code'),
    ...getSelectedValues('mediaFunnel').map(() => 'Media Funnel'),
    ...getSelectedValues('adProduct').map(() => 'Ad Product'),
    ...getSelectedValues('biddingGoal').map(() => 'Bidding Goal'),
    ...getSelectedValues('url').map(() => 'URL')
  ];
  const customVal = document.getElementById('custom').value.trim();
  if (customVal) labels.push('Custom');

  let html = '';
  parts.forEach((p, i) => {
    html += `<div class="breakdown-item">
      <span class="bd-label">${labels[i] || '?'}</span>
      <span class="bd-value">${p}</span>
    </div>`;
  });
  el.innerHTML = html;
  el.style.display = 'block';
}

/* ═══════════════════════════════════════════════════════════════════════════
   COPY & HISTORY
   ═════════════════════════════════════════════════════════════════════════ */
function copyResult() {
  const text = document.querySelector('.preview-box .result')?.textContent;
  if (!text) return alert('Nothing to copy. Generate a name first.');
  navigator.clipboard.writeText(text).then(() => alert('✅ Copied: ' + text));
}

function copyAll() {
  const text = document.querySelector('.preview-box .result')?.textContent;
  if (!text) return alert('Nothing to copy.');
  const line = `${text},`;
  navigator.clipboard.writeText(line).then(() => alert('✅ Copied (CSV format): ' + line));
}

function addHistory(name) {
  const now = new Date().toLocaleTimeString();
  copyHistory.unshift({ name, time: now });
  if (copyHistory.length > 20) copyHistory.pop();
  renderHistory();
}

function renderHistory() {
  const el = document.getElementById('historyList');
  if (copyHistory.length === 0) {
    el.innerHTML = '<div style="padding:20px;text-align:center;color:var(--text2);font-size:13px;">No history yet</div>';
    return;
  }
  el.innerHTML = copyHistory.map(h => `
    <div class="history-item">
      <span class="history-text" onclick="copyText('${h.name}')" title="Click to copy">${h.name}</span>
      <span class="history-time">${h.time}</span>
    </div>`).join('');
}

function copyText(t) {
  navigator.clipboard.writeText(t).then(() => alert('✅ Copied: ' + t));
}

/* ═══════════════════════════════════════════════════════════════════════════
   CUSTOM HELP TAGS
   ═════════════════════════════════════════════════════════════════════════ */
function renderCustomHelpTags() {
  const container = document.getElementById('customHelpTags');
  (ENUM_DATA.custom_ref || []).forEach(tag => {
    const el = document.createElement('span');
    el.className = 'help-tag';
    el.textContent = tag;
    el.onclick = () => toggleCustomTag(tag, el);
    container.appendChild(el);
  });
}

function toggleCustomTag(tag, el) {
  const input = document.getElementById('custom');
  const current = input.value.trim().toLowerCase();
  const tags = current ? current.split(' ').filter(Boolean) : [];
  if (tags.includes(tag)) {
    el.classList.remove('selected');
    input.value = tags.filter(t => t !== tag).join(' ');
  } else {
    el.classList.add('selected');
    tags.push(tag);
    input.value = tags.join(' ');
  }
}

function toggleCustomHelp() {
  const body = document.getElementById('customHelpBody');
  const arrow = document.getElementById('customHelpArrow');
  body.classList.toggle('hidden');
  arrow.textContent = body.classList.contains('hidden') ? '▶' : '▼';
}

/* ═══════════════════════════════════════════════════════════════════════════
   RESET
   ═════════════════════════════════════════════════════════════════════════ */
function resetAll() {
  document.querySelectorAll('select').forEach(s => {
    if (s.multiple) {
      Array.from(s.options).forEach(o => o.selected = false);
    } else {
      s.selectedIndex = 0;
    }
  });
  document.getElementById('custom').value = '';
  document.getElementById('previewBox').innerHTML =
    '<span style="color:var(--text2);">Fill in required fields and click "Generate Name"</span>';
  document.getElementById('previewBox').dataset.generated = '';
  document.getElementById('previewActions').style.display = 'none';
  document.getElementById('breakdown').style.display = 'none';
  document.getElementById('validationWarning').classList.remove('show');
  /* Reset custom help tags visual */
  document.querySelectorAll('.help-tag').forEach(t => t.classList.remove('selected'));
  /* Reset product code options */
  updateProductCodeOptions();
}

/* ═══════════════════════════════════════════════════════════════════════════
   BOOT
   ═════════════════════════════════════════════════════════════════════════ */
fetchData();
</script>
</body>
</html>"""


# ── 主程序 ───────────────────────────────────────────────────────────────────
if __name__ == "__main__":
    print("============================")
    print("  Ad Campaign Name Builder - Generator v7")
    print("  Product Code Map + Funnel-Goal + Compact")
    print("============================")

    data = read_excel_data()
    generate_data_json(data)
    generate_product_code_map_json()

    html = get_html_template()
    # Embed product code map as JS fallback
    map_json_str = json.dumps(PRODUCT_CODE_MAP, ensure_ascii=False, indent=2)
    html = html.replace("##PRODUCT_CODE_MAP_PLACEHOLDER##", map_json_str)

    with open(HTML_OUTPUT, "w", encoding="utf-8") as f:
        f.write(html)
    print("OK: Generated index.html")
    print("============================")
    print("Output files:")
    print("  - " + DATA_JSON)
    print("  - " + MAP_JSON)
    print("  - " + HTML_OUTPUT)
    print("============================")
